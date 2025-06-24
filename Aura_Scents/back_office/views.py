
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.http import JsonResponse

from .forms import *

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.core.paginator import Paginator
from django.db.models import Q

from django.views.decorators.http import require_POST

from django.shortcuts import render, redirect, get_object_or_404
from .models import *

from uuid import uuid4
import os
from django.core.exceptions import ValidationError

from store.models import *

from django.db import transaction

from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.db.models import Sum, Count



#Admin View
def admin_login(request):
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            return redirect('admin_dashboard')
    else:
        form = UserLoginForm()
    return render(request, 'back_office/admin_login.html', {'form': form})

# Dashboard View
@staff_member_required(login_url='admin_login') 
def admin_dashboard(request):
    return render(request, 'back_office/admin_dashboard.html')


# Log Out View
def admin_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('admin_login')


# User List and Search  
@staff_member_required(login_url='admin_login') 
def user_list(request):
    query = request.GET.get('q')
    users = User.objects.filter(is_superuser=False)

    if query:
        users = users.filter(Q(username__icontains=query) | Q(email__icontains=query))

    users = users.order_by('-id')  # Latest first
    paginator = Paginator(users, 5)
    page = request.GET.get('page')
    users = paginator.get_page(page)

    return render(request, 'back_office/user_list.html', {'users': users, 'query': query})

# Block/Unblock Users
@require_POST
def toggle_block_user(request, user_id):
    user = User.objects.get(id=user_id)
    user.is_blocked = not user.is_blocked
    user.save()
    messages.success(request, f"User {'blocked' if user.is_blocked else 'unblocked'} successfully.")
    return redirect('user_list')


# Category List
@staff_member_required(login_url='admin_login') 
def category_list(request):
    query = request.GET.get('q')
    categories = Category.objects.filter(is_deleted=False)

    if query:
        categories = categories.filter(name__icontains=query)

    categories = categories.order_by('-created_at')
    paginator = Paginator(categories, 5)
    page = request.GET.get('page')
    categories = paginator.get_page(page)

    return render(request, 'back_office/category_list.html', {'categories': categories, 'query': query})

# Category Add
@staff_member_required(login_url='admin_login') 
def category_add(request):
    if request.method == 'POST':
        name = request.POST['name']
        is_blocked = not ('category-status' in request.POST)
        Category.objects.create(name=(name), is_blocked=is_blocked)
        return redirect('category_list')
    return render(request, 'back_office/category_add.html')

#Category Edit
@staff_member_required(login_url='admin_login') 
def category_edit(request, category_id):
    category = Category.objects.get(id=category_id)
    if request.method == 'POST':
        category.name = request.POST['name']
        category.is_blocked = not ('category-status' in request.POST)       
        category.save()
        return redirect('category_list')
    return render(request, 'back_office/category_edit.html', {'category': category})

#Category Delete
@staff_member_required(login_url='admin_login') 
def category_delete(request, category_id):
    category = Category.objects.get(id=category_id)
    category.is_deleted = True
    category.save()
    return redirect('category_list')

#Product List
@staff_member_required(login_url='admin_login') 
def product_list(request):
    query = request.GET.get('q', '')
    products = Product.objects.filter(is_deleted=False)

    if query:
        products = products.filter(name__icontains=query)

    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'back_office/product_list.html', {'page_obj': page_obj, 'query':query} )





#Product Manage
@staff_member_required(login_url='admin_login')
def product_manage(request, product_id=None):
    # Determine if this is an add or edit operation
    is_edit = product_id is not None
    
    if is_edit:
        product = get_object_or_404(Product, id=product_id)
        variants = ProductVariant.objects.filter(product=product)
        images = ProductImage.objects.filter(product=product)
    else:
        product = None
        variants = []
        images = []

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        price = request.POST.get('price')
        stock = request.POST.get('stock')
        category_id = request.POST.get('category')
        is_blocked = not ('product-status' in request.POST)
        new_images = request.FILES.getlist('images')
        use_variants = request.POST.get('use_variants') == 'on'
        
        # Handle image removals in edit mode
        removed_images_str = request.POST.get('removed_images', '')
        removed_image_ids = []
        if removed_images_str:
            try:
                removed_image_ids = [int(id.strip()) for id in removed_images_str.split(',') if id.strip()]
            except ValueError:
                pass  # Invalid format, ignore

        # Validation
        if not name or not description or not category_id:
            error = "Name, description and category are required fields."
            categories = Category.objects.filter(is_deleted=False, is_blocked=False)
            return render(request, 'back_office/product_manage.html', locals())

        # Image validation - check total images after removals and additions
        if is_edit:
            # Count existing images that won't be removed
            existing_image_count = ProductImage.objects.filter(product=product).exclude(id__in=removed_image_ids).count()
            total_images = existing_image_count + len(new_images)
        else:
            total_images = len(new_images)

        if total_images < 3:
            error = f"At least 3 images are required. You currently have {total_images} image(s)."
            categories = Category.objects.filter(is_deleted=False, is_blocked=False)
            return render(request, 'back_office/product_manage.html', locals())

        if not use_variants:
            try:
                price = float(price)
                stock = int(stock)
                if price < 0 or stock < 0:
                    raise ValidationError("Price and stock must be non-negative.")
            except (ValueError, TypeError):
                error = "Invalid price or stock format."
                categories = Category.objects.filter(is_deleted=False, is_blocked=False)
                return render(request, 'back_office/product_manage.html', locals())
        else:
            price = 0  # Placeholder
            stock = 0  # Will sum from variants

        category = get_object_or_404(Category, id=category_id)

        # Create or Update Product
        if is_edit:
            product.name = name
            product.description = description
            product.price = price
            product.stock = stock
            product.is_blocked = is_blocked
            product.category = category
            product.save()
        else:
            product = Product.objects.create(
                name=name,
                description=description,
                price=price,
                stock=stock,
                is_blocked=is_blocked,
                category=category
            )

        # Handle Image Removals (for edit mode)
        if is_edit and removed_image_ids:
            try:
                # Get the images to be removed
                images_to_remove = ProductImage.objects.filter(
                    product=product, 
                    id__in=removed_image_ids
                )
                
                # Delete the actual image files from storage (optional)
                for img in images_to_remove:
                    if img.image and hasattr(img.image, 'delete'):
                        img.image.delete(save=False)
                
                # Delete the database records
                images_to_remove.delete()
                
            except Exception as e:
                # Log the error if you have logging set up
                print(f"Error removing images: {e}")

        # Handle New Image Additions
        if new_images:
            for image in new_images:
                try:
                    ProductImage.objects.create(product=product, image=image)
                except Exception as e:
                    # Log the error if you have logging set up
                    print(f"Error adding image: {e}")

        # Handle Variants
        if use_variants:
            variant_volumes = request.POST.getlist('variant_volume[]')
            variant_units = request.POST.getlist('variant_unit[]')
            variant_prices = request.POST.getlist('variant_price[]')
            variant_stocks = request.POST.getlist('variant_stock[]')
            variant_ids = request.POST.getlist('variant_id[]')

            # Validate variant data
            if not variant_volumes or not variant_units or not variant_prices or not variant_stocks:
                error = "All variant fields are required when using variants."
                categories = Category.objects.filter(is_deleted=False, is_blocked=False)
                return render(request, 'back_office/product_manage.html', locals())

            # First delete any variants that were removed
            if is_edit:
                existing_variant_ids = [str(v.id) for v in variants]
                variants_to_remove = [v_id for v_id in existing_variant_ids if v_id not in variant_ids]
                if variants_to_remove:
                    ProductVariant.objects.filter(id__in=variants_to_remove).delete()

            total_stock = 0
            valid_variants_count = 0
            
            for idx, (vol, unit, pr, stk) in enumerate(zip(variant_volumes, variant_units, variant_prices, variant_stocks)):
                if not vol or not unit or not pr or not stk:
                    continue  # Skip incomplete rows

                try:
                    vol = float(vol)
                    pr = float(pr)
                    stk = int(stk)
                    
                    if vol <= 0 or pr < 0 or stk < 0:
                        error = "Variant volume must be greater than 0, price and stock must be non-negative."
                        categories = Category.objects.filter(is_deleted=False, is_blocked=False)
                        return render(request, 'back_office/product_manage.html', locals())
                    
                    total_stock += stk
                    valid_variants_count += 1
                    
                except (ValueError, TypeError):
                    error = "Invalid variant data format."
                    categories = Category.objects.filter(is_deleted=False, is_blocked=False)
                    return render(request, 'back_office/product_manage.html', locals())

                variant_id = variant_ids[idx] if idx < len(variant_ids) and variant_ids[idx] else None
                
                if variant_id:  # Update existing variant
                    try:
                        variant = ProductVariant.objects.get(id=variant_id, product=product)
                        variant.volume = vol
                        variant.unit = unit
                        variant.price = pr
                        variant.stock = stk
                        variant.save()
                    except ProductVariant.DoesNotExist:
                        # Create new variant if the existing one doesn't exist
                        ProductVariant.objects.create(
                            product=product,
                            volume=vol,
                            unit=unit,
                            price=pr,
                            stock=stk
                        )
                else:  # Create new variant
                    ProductVariant.objects.create(
                        product=product,
                        volume=vol,
                        unit=unit,
                        price=pr,
                        stock=stk
                    )
            
            # Ensure at least one valid variant exists
            if valid_variants_count == 0:
                error = "At least one valid variant is required when using variants."
                categories = Category.objects.filter(is_deleted=False, is_blocked=False)
                return render(request, 'back_office/product_manage.html', locals())
            
            # Update product stock to sum of all variants
            product.stock = total_stock
            product.save()
        else:
            # If not using variants, delete any existing variants
            if is_edit:
                ProductVariant.objects.filter(product=product).delete()

        
        return redirect('product_list')

    categories = Category.objects.filter(is_deleted=False, is_blocked=False)
    return render(request, 'back_office/product_manage.html', {
        'categories': categories,
        'product': product,
        'variants': variants,
        'images': images,
        'is_edit': is_edit
    })



# Product Edit
@staff_member_required(login_url='admin_login') 
def product_edit(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.filter(is_deleted=False)

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        price = request.POST.get('price')
        stock = request.POST.get('stock')
        category_id = request.POST.get('category')
        is_blocked =  not ('product-status' in request.POST)
        images = request.FILES.getlist('images')

        if not name or not description or not price or not stock or not category_id:
            error = "All fields are required."
            return render(request, 'back_office/product_edit.html', {
                'error': error,
                'product': product,
                'edit': True,
                'form_data': request.POST,
                'categories': categories
            })

        product.name = name
        product.description = description
        product.price = price
        product.stock = stock
        product.category_id = category_id
        product.is_blocked = is_blocked
        product.save()

        for image in images:
            ProductImage.objects.create(product=product, image=image)

        return redirect('product_list')

    return render(request, 'back_office/product_edit.html', {
        'product': product,
        'edit': True,
        'categories': categories,
        'is_blocked': product.is_blocked,
    })


# Product Delete
@staff_member_required(login_url='admin_login') 
def product_delete(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product.is_deleted = True
    product.save()
    return redirect('product_list')

def order_list(request):
    orders = Order.objects.select_related('user').all()  # Removed incorrect variant reference

    # === SEARCH ===
    query = request.GET.get('q')
    if query:
        orders = orders.filter(
            Q(order_id__icontains=query) |
            Q(user__email__icontains=query) |
            Q(user__username__icontains=query)
        )

    # === FILTER BY STATUS ===
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status=status_filter)

    # === SORTING ===
    sort_by = request.GET.get('sort')
    if sort_by == 'date_asc':
        orders = orders.order_by('created_at')
    else:  # default or date_desc
        orders = orders.order_by('-created_at')

    # === PAGINATION ===
    paginator = Paginator(orders, 10)  # 10 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'orders': page_obj,
        'query': query,
        'status_filter': status_filter,
        'sort_by': sort_by,
    }
    return render(request, 'back_office/order_list.html', context)



def order_detail(request, order_id):
    order = get_object_or_404(
        Order.objects.prefetch_related('items__product__images', 'items__product__variants'),
        id=order_id
    )
    return render(request, 'back_office/order_detail.html', {'order': order})



import logging

logger = logging.getLogger(__name__)

@staff_member_required(login_url='admin_login')
def update_order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        item_id = request.POST.get('item_id')  # Optional: for item-level updates
        old_status = order.status

        with transaction.atomic():
            if item_id:  # Item-level status update
                order_item = get_object_or_404(OrderItem, id=item_id, order=order)
                old_item_status = order_item.status

                if new_status != old_item_status:
                    order_item.status = new_status
                    order_item.save()

                    # Handle item-level refunds
                    if new_status in ['Cancelled', 'Returned'] and (order.is_paid or order.payment_method != 'COD'):
                        wallet, _ = Wallet.objects.get_or_create(user=order.user)
                        refund_amount = order_item.subtotal()
                        WalletTransaction.objects.create(
                            wallet=wallet,
                            order=order,
                            transaction_type='credit',
                            amount=refund_amount,
                            description=f"Refund for {order_item.product.name} in order {order.order_id} ({new_status})"
                        )
                        wallet.balance += refund_amount
                        wallet.save()

                    # Update product stock
                    order_item.product.stock += order_item.quantity
                    order_item.product.save()

                # Update order status based on item statuses
                if all(item.status == 'Cancelled' for item in order.items.all()):
                    order.status = 'Cancelled'
                    order.refund_processed = True
                elif any(item.status == 'Return Requested' for item in order.items.all()):
                    order.status = 'Return Requested'
                order.save()

            else:  # Order-level status update
                if new_status != old_status:
                    order.status = new_status
                    if new_status == 'Delivered' and order.payment_method == 'COD':
                        order.is_paid = True
                    order.save()

                # Handle refunds for order-level cancellations/returns
                if new_status in ['Cancelled', 'Returned'] and not order.refund_processed:
                    wallet, _ = Wallet.objects.get_or_create(user=order.user)
                    refund_amount = order.total_amount if (order.is_paid or new_status == 'Cancelled') else Decimal('0.00')
                    if refund_amount > 0:
                        WalletTransaction.objects.create(
                            wallet=wallet,
                            order=order,
                            transaction_type='credit',
                            amount=refund_amount,
                            description=f"Refund for order {order.order_id} ({new_status})"
                        )
                        wallet.balance += refund_amount
                        wallet.save()
                    order.refund_processed = True
                    order.save()

                # Handle inventory for cancellations/returns
                if new_status in ['Cancelled', 'Returned']:
                    for item in order.items.all():
                        product = item.product
                        product.stock += item.quantity
                        product.save()
                        item.status = new_status  # Update item status to match order
                        item.save()

            messages.success(request, f'Order status updated to {new_status}')
            return redirect('order_details', order_id=order.id)
    return redirect('order_list')


# Coupon View
@login_required
def manage_coupons(request):
    """View to manage coupons (create and edit)"""
    coupons = Coupon.objects.all()

    # Check if editing a coupon
    edit_coupon_id = request.GET.get('edit')
    if edit_coupon_id:
        coupon = get_object_or_404(Coupon, id=edit_coupon_id)
        form = CouponForm(request.POST or None, instance=coupon)
    else:
        form = CouponForm(request.POST or None)

    if request.method == 'POST' and 'create_coupon' in request.POST:
        if form.is_valid():
            form.save()
            messages.success(request, f"Coupon {'updated' if edit_coupon_id else 'created'} successfully!")
            return redirect('manage_coupons')
        else:
            messages.error(request, "Please correct the errors below.")

    return render(request, 'back_office/manage_coupons.html', {'form': form, 'coupons': coupons})

@login_required
def delete_coupon(request, coupon_id):
    """View to delete a specific coupon"""
    if request.method == 'POST':
        coupon = get_object_or_404(Coupon, id=coupon_id)
        coupon_code = coupon.code  # Store coupon code for response
        try:
            coupon.delete()
            message = f"Coupon '{coupon_code}' deleted successfully!"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': message
                })
            messages.success(request, message)
            return redirect('manage_coupons')
        except Exception as e:
            message = f"Failed to delete coupon '{coupon_code}'. Error: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': message
                }, status=500)
            messages.error(request, message)
            return redirect('manage_coupons')
    else:
        messages.error(request, "Invalid request method.")
        return redirect('manage_coupons')
    

def manage_offers(request):
    # Use select_related for Offer and prefetch_related for product_offer and category_offer
    offers = Offer.objects.all().select_related().prefetch_related('product_offer__product', 'category_offer__category')
    edit_offer = None
    offer_form = OfferForm()
    product_offer_form = ProductOfferForm()
    category_offer_form = CategoryOfferForm()

    if request.GET.get('edit'):
        edit_offer = get_object_or_404(Offer, id=request.GET.get('edit'))
        offer_form = OfferForm(instance=edit_offer)
        if edit_offer.offer_type == 'product':
            try:
                product_offer = edit_offer.product_offer
                product_offer_form = ProductOfferForm(instance=product_offer)
            except ProductOffer.DoesNotExist:
                product_offer_form = ProductOfferForm()
        elif edit_offer.offer_type == 'category':
            try:
                category_offer = edit_offer.category_offer
                category_offer_form = CategoryOfferForm(instance=category_offer)
            except CategoryOffer.DoesNotExist:
                category_offer_form = CategoryOfferForm()

    if request.method == 'POST' and 'create_offer' in request.POST:
        offer_form = OfferForm(request.POST, instance=edit_offer)
        product_offer_form = ProductOfferForm(request.POST)
        category_offer_form = CategoryOfferForm(request.POST)

        is_valid = True

        # Validate offer form
        if not offer_form.is_valid():
            is_valid = False

        # Validate based on offer type
        offer_type = offer_form.cleaned_data.get('offer_type') if offer_form.is_valid() else request.POST.get('offer_type')
        if offer_type == 'product':
            if not product_offer_form.is_valid() or not product_offer_form.cleaned_data.get('product'):
                is_valid = False
                product_offer_form.add_error('product', 'This field is required.')
        elif offer_type == 'category':
            if not category_offer_form.is_valid() or not category_offer_form.cleaned_data.get('category'):
                is_valid = False
                category_offer_form.add_error('category', 'This field is required.')

        if is_valid:
            offer = offer_form.save()

            if offer.offer_type == 'product':
                # Delete any existing CategoryOffer
                CategoryOffer.objects.filter(offer=offer).delete()
                try:
                    # Check if a ProductOffer already exists
                    product_offer = ProductOffer.objects.get(offer=offer)
                    # Update existing ProductOffer
                    product_offer.product = product_offer_form.cleaned_data['product']
                    product_offer.save()
                except ProductOffer.DoesNotExist:
                    # Create new ProductOffer
                    product_offer = product_offer_form.save(commit=False)
                    product_offer.offer = offer
                    product_offer.save()
                messages.success(request, 'Product offer saved successfully!')
                return redirect('manage_offers')
            elif offer.offer_type == 'category':
                # Delete any existing ProductOffer
                ProductOffer.objects.filter(offer=offer).delete()
                try:
                    # Check if a CategoryOffer already exists
                    category_offer = CategoryOffer.objects.get(offer=offer)
                    # Update existing CategoryOffer
                    category_offer.category = category_offer_form.cleaned_data['category']
                    category_offer.save()
                except CategoryOffer.DoesNotExist:
                    # Create new CategoryOffer
                    category_offer = category_offer_form.save(commit=False)
                    category_offer.offer = offer
                    category_offer.save()
                messages.success(request, 'Category offer saved successfully!')
                return redirect('manage_offers')
        else:
            if offer_form.is_valid() and not edit_offer:
                offer = offer_form.save(commit=False)
                offer.delete() # Clean up if invalid and new offer
            messages.error(request, 'Invalid offer details.')

    return render(request, 'back_office/manage_offers.html', {
        'offers': offers,
        'offer_form': offer_form,
        'product_offer_form': product_offer_form,
        'category_offer_form': category_offer_form,
        'edit_offer': edit_offer,
    })

@require_POST
def delete_offer(request, offer_id):
    offer = get_object_or_404(Offer, id=offer_id)
    offer_name = offer.name
    offer.delete()
    return JsonResponse({
        'success': True,
        'message': f'Offer "{offer_name}" deleted successfully!'
    })

def sales_report(request):
    # Default filter: last 30 days
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    filter_type = request.GET.get('filter_type', 'custom')

    # Handle preset filters
    if filter_type == 'daily':
        start_date = end_date
    elif filter_type == 'weekly':
        start_date = end_date - timedelta(days=7)
    elif filter_type == 'monthly':
        start_date = end_date - timedelta(days=30)
    elif filter_type == 'yearly':
        start_date = end_date - timedelta(days=365)

    # Handle custom date range
    if request.GET.get('start_date'):
        try:
            start_date = timezone.datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        except ValueError:
            pass
    if request.GET.get('end_date'):
        try:
            end_date = timezone.datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
        except ValueError:
            pass

    # Query orders within date range
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    )

    # Calculate metrics
    total_sales_count = orders.count()
    total_order_amount = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_discount = sum(order.discount for order in orders) or Decimal('0.00')
    net_sales = total_order_amount - total_discount

    # Handle export
    export_format = request.GET.get('export')
    if export_format:
        context = {
            'start_date': start_date,
            'end_date': end_date,
            'total_sales_count': total_sales_count,
            'total_order_amount': total_order_amount,
            'total_discount': total_discount,
            'net_sales': net_sales,
            'orders': orders,
            'request': request,  # Add request to context for PDF rendering
        }

        if export_format == 'pdf':
            template = get_template('back_office/sales_report_pdf.html')
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.pdf"'
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return HttpResponse('Error generating PDF', status=500)
            return response

        elif export_format == 'excel':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sales Report'

            # Headers
            headers = ['Order ID', 'Date', 'Customer', 'Total Amount', 'Discount', 'Net Amount']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(bottom=Side(style='thin'))

            # Data
            for row, order in enumerate(orders, 2):
                worksheet.cell(row=row, column=1).value = order.order_id
                worksheet.cell(row=row, column=2).value = order.created_at.date()
                worksheet.cell(row=row, column=3).value = order.user.email
                worksheet.cell(row=row, column=4).value = float(order.total_amount)
                worksheet.cell(row=row, column=5).value = float(order.discount)
                worksheet.cell(row=row, column=6).value = float(order.total_amount - order.discount)

            # Summary
            summary_row = len(orders) + 3
            worksheet.cell(row=summary_row, column=1).value = 'Total'
            worksheet.cell(row=summary_row, column=4).value = float(total_order_amount)
            worksheet.cell(row=summary_row, column=5).value = float(total_discount)
            worksheet.cell(row=summary_row, column=6).value = float(net_sales)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'
            workbook.save(response)
            return response

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
        'total_sales_count': total_sales_count,
        'total_order_amount': total_order_amount,
        'total_discount': total_discount,
        'net_sales': net_sales,
        'orders': orders,
    }

    return render(request, 'back_office/sales_report.html', context)

@require_POST
@staff_member_required(login_url='admin_login')
def process_refund(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if order.refund_processed:
        return JsonResponse({'success': False, 'message': 'Refund already processed'})
    
    if order.status not in ['Cancelled', 'Returned'] or (order.payment_method == 'COD' and not order.is_paid):
        return JsonResponse({'success': False, 'message': 'Refund not applicable for this order'})
    
    try:
        with transaction.atomic():
            wallet, _ = Wallet.objects.get_or_create(user=order.user)
            refund_amount = order.total_amount
            WalletTransaction.objects.create(
                wallet=wallet,
                order=order,
                transaction_type='credit',
                amount=refund_amount,
                description=f"Manual refund for order {order.order_id}"
            )
            wallet.balance += refund_amount
            wallet.save()
            order.refund_processed = True
            order.save()
            logger.info(f"Manual refund of ₹{refund_amount} processed for order {order.order_id}")
        return JsonResponse({'success': True, 'message': f'Refunded ₹{refund_amount} to wallet'})
    except Exception as e:
        logger.error(f"Error processing refund for order {order.order_id}: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error processing refund: {str(e)}'})




